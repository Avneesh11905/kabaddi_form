
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from app.config import settings

def get_slots():
    try:
        # Check if file exists in the root directory relative to where CWD will be
        # Assuming run from project root
        if not os.path.exists(settings.INPUT_FILE):
             # Fallback/Debug
             print(f"File not found: {settings.INPUT_FILE} at {os.getcwd()}")
             return ["Slot A", "Slot B", "Slot C"]
             
        df = pd.read_excel(settings.INPUT_FILE)
        slots = df.iloc[0].tolist() 
        return [str(s) for s in slots if pd.notna(s)]
    except Exception as e:
        print(f"Error reading slots: {e}")
        return ["Slot A", "Slot B", "Slot C"]

def generate_excel_bytes(submissions, slots, target_date):
    data_dict = {slot: [] for slot in slots}
    
    for sub in submissions:
        reg = sub.get("reg_no")
        start_slots = sub.get("slots", [])
        for slot in start_slots:
            if slot in data_dict:
                data_dict[slot].append(reg)
                
    if data_dict:
        max_len = max(len(l) for l in data_dict.values()) if any(data_dict.values()) else 0
        for col in data_dict:
            while len(data_dict[col]) < max_len:
                data_dict[col].append(None)
    
    new_df = pd.DataFrame(data_dict)
    
    output = io.BytesIO()
    
    title_text = f"Kabaddi {target_date.day} {target_date.strftime('%b')} {target_date.year}"
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        new_df.to_excel(writer, index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        worksheet['A1'] = title_text
        if len(slots) > 1:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(slots))
        
        title_cell = worksheet['A1']
        title_cell.font = Font(size=12, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        header_font = Font(bold=True)
        for cell in worksheet[2]:
            cell.font = header_font
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    output.seek(0)
    return output
